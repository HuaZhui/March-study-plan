# -*- coding: utf-8 -*-
"""重建三个月多方向学习计划：后端、Android、H5、Cocos、Unity3D、运维"""

import os
from datetime import date, timedelta
from openpyxl import load_workbook

DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DIR, "三个月全栈入门计划.xlsx")
# 66 个工作日 · 6/1 起 · 仅周一至周五 · 8/31 结业（12 内容周合并自原 84 天）
START = date(2026, 6, 1)
END_DEADLINE = date(2026, 8, 31)
TOTAL_LEARNING_DAYS = 66
WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

# 每周原 7 天按组合并（0-based 索引）；合计 66 天
WEEK_MERGE_GROUPS = {
    "W1": [[0], [1], [2], [3, 4], [5], [6]],           # 6：HTML+CSS 读码合并
    "W2": [[0], [1], [2], [3], [4], [5, 6]],           # 6：AI 工作流+验收合并
    "W3": [[0], [1, 2], [3], [4, 5], [6]],            # 5
    "W4": [[0], [1, 2], [3], [4], [5], [6]],           # 6：SQL 三天压成两天
    "W5": [[0], [1, 2, 3], [4], [5], [6]],             # 5：UI+列表合并
    "W6": [[0], [1], [2, 3], [4], [5], [6]],           # 6
    "W7": [[0, 1], [2], [3], [4, 5], [6]],             # 5
    "W8": [[0], [1, 2], [3], [4, 5], [6]],             # 5
    "W9": [[0], [1, 2], [3], [4], [5], [6]],           # 6：Redis 两天合并
    "W10": [[0], [1], [2], [3, 4], [5], [6]],          # 6
    "W11": [[0], [1, 2], [3], [4, 5], [6]],            # 5
    "W12": [[0, 1], [2], [3], [4], [5, 6]],            # 5
}
MILESTONE_WEEKS = {"M1": "W4", "M2": "W8", "M3": "W12"}


def learning_date(start, index):
    """第 index 个学习日（0-based），仅周一至周五，跳过周末。"""
    d = start
    while d.weekday() >= 5:
        d += timedelta(days=1)
    for _ in range(index):
        d += timedelta(days=1)
        while d.weekday() >= 5:
            d += timedelta(days=1)
    return d
LEARN_LOOP = "看(15~30min)→敲→跑→改；上午看文档下午做交付物；卡住：报错→查官方文档→对比示例→求助"
H5_LEARN_LOOP = "懂 H5 在干什么→写清需求→AI 实现→浏览器/真机验收；卡住：F12 Console/Network→改 prompt→对照文档"
CODE_ROOT = r"e:\三个月速成\learning-code"
REVIEW = "写3句：学了啥/卡在哪/明天干啥"

H5_GOALS = [
    "能看懂公司 H5 项目基本代码结构与目录",
    "理解页面打开→调接口→渲染数据的完整链路",
    "理解 H5 与后端、Native 的协作边界",
    "理解路由、组件、API 封装、状态管理、移动端样式适配",
    "能在浏览器/真机做基本验收，会用 F12 辅助定位问题",
    "能对 AI 生成的 H5 代码做基本审核，识别明显风险",
    "会用团队 AI Agent/Skill 完成原型转文档、页面实现、E2E、部署等常见任务",
    "能根据自己的基础制定后续学习计划",
]

H5_PHILOSOPHY = "不必先系统学完 HTML/CSS/JS 再开工——先懂 H5 在干什么、能把需求说清楚，实现交给 AI，你在浏览器和真机上验收结果。"

OPS_GOALS = [
    "了解不同运维岗位的职责边界和协作方式",
    "掌握常见基础设施、网络、安全、发布、数据库相关概念",
    "能够理解日常故障沟通中的专业词汇",
    "对常见业务故障具备初步定位思路",
    "通过自查题发现知识薄弱点，便于后续深入学习",
]

OPS_LEARN_LOOP = "先懂岗位边界与核心概念→记术语能听懂→练故障定位思路→自查薄弱点；动手以验证概念为主（Docker/Nginx 各练一次），不必先考运维证书"
OPS_PHILOSOPHY = "运维学习重点是「听得懂、说得清、会协作、能初判」——先建立认知地图和故障沟通能力，再按自查结果决定后续深入方向。"

# (模块, 今日目标, 交付物, 阅读, 文档名, 文档链接, 教程, 怎么创建, 文件夹, 上午, 下午, 验收, 学时)
WEEKS = {
    "W1": {
        "阶段": "第1月·工具与读码",
        "days": [
            ("环境", "安装 VS Code/Git/Postman/Chrome；对齐 H5 八条学习目标", "工具可用+目标对照表", "H5 学习目标（8条）", "Chrome DevTools", "https://developer.chrome.com/docs/devtools", "Cursor 使用入门 | 团队 AI Skill 说明", f"建 {CODE_ROOT}；读 学习产出/web学习产出.html", "day01-intro", "装 VS Code/Git/Postman/Chrome", "抄写8条H5目标+核心理念；git init", "☐ 工具可用 ☐ 8目标能口述", 7),
            ("Git", "push/pull；远程仓库", "代码推到 Gitee/GitHub", "Pro Git 第2章", "Pro Git", "https://git-scm.com/book/zh/v2", "菜鸟 Git", f"新建 day02-git", "day02-git", "注册远程仓；git remote", "push/pull 验证", "☐ 远程可见 ☐ 会 push/pull", 7),
            ("MySQL", "安装 MySQL 或 Docker（懂后端协作）", "能连上数据库", "MySQL 安装向导", "MySQL 文档", "https://dev.mysql.com/doc/", "Docker 装 MySQL", "—", "—", "安装 MySQL8 或 docker mysql", "DBeaver 连接；SHOW DATABASES", "☐ 客户端连上 ☐ 知道 H5 不直连库", 7),
            ("读码", "读懂 HTML 页面结构（不从头写）", "标注页面结构笔记", "MDN：HTML 入门（只读）", "MDN HTML", "https://developer.mozilla.org/zh-CN/docs/Learn/HTML", "打开 学习产出/六一儿童节.html", f"阅读 e:\\三个月速成\\学习产出\\", "h5-read-html", "F12 Elements 看 DOM 树", "标注 header/section/脚本位置", "☐ 能指认主要区块 ☐ 知道哪是结构哪是脚本", 7),
            ("读码", "读懂 CSS 布局与移动适配", "viewport/rem 对照笔记", "MDN：viewport/Flex（只读）", "MDN CSS", "https://developer.mozilla.org/zh-CN/docs/Learn/CSS", "B站：H5 rem 适配（只看对照）", "h5-read-css", "h5-read-css", "找 viewport/meta 与 rem 设置", "手机模式预览；记录适配方式", "☐ 能说明如何适配手机宽 ☐ 能找布局代码", 7),
            ("读码", "读懂 JS 与 Network 请求链路", "一条 API 链路图", "MDN：Network 面板", "Chrome DevTools", "https://developer.chrome.com/docs/devtools/network", "—", "h5-read-network", "h5-read-network", "F12 Network 刷新页面", "跟一条 XHR：请求→响应→页面变化", "☐ 能画链路图 ☐ 会看 Status/Preview", 7),
            ("验收", "W1 工具与读码验收", "F12+Git+MySQL+读码笔记", "—", "Chrome DevTools", "https://developer.chrome.com/docs/devtools", "—", "—", "day07-check", "自测：Elements/Network/Console", "git push；预习 W2 H5 专周", "☐ W1全过", 6),
        ],
    },
    "W2": {
        "阶段": "第1月·H5验收向",
        "days": [
            ("H5", "【目标1】看懂 H5 项目目录与代码结构", "项目目录说明文档", "典型 H5 目录规范", "Vite+Vue H5 模板", "https://cn.vuejs.org/guide/quick-start.html", "npm create vite@latest h5-demo", f"npm create vite@latest h5-demo → {CODE_ROOT}\\h5-demo", "h5-demo", "梳理 src/router/api/views 等目录", "对照 学习产出 说明各文件夹职责", "☐ 能讲清目录 ☐ 项目能 npm run dev", 7),
            ("H5", "【目标2】页面打开→调接口→渲染 完整链路", "数据流时序图/笔记", "Axios 请求流程", "MDN Fetch/XHR", "https://developer.mozilla.org/zh-CN/docs/Web/API/Fetch_API", "jsonplaceholder 对照", "h5-demo", "h5-demo", "让 AI 加列表页+mock 接口", "F12 Network 逐步跟：DOM→JS→请求→渲染", "☐ 链路能口述 ☐ Network 能对上", 7),
            ("H5", "【目标3】H5/后端/Native 协作边界", "协作边界一页纸", "前后端职责划分", "MDN CORS（了解）", "https://developer.mozilla.org/zh-CN/docs/Web/HTTP/CORS", "微信 JS-SDK 文档（Native 桥）", "h5-notes", "h5-notes", "画 H5/后端/Native 分工表", "写：谁提供接口/谁渲染/谁调原生", "☐ 边界文档完成 ☐ 能答联调谁改啥", 7),
            ("H5", "【目标4】路由/组件/API封装/状态/移动适配", "概念对照表（5项）", "Vue Router/Pinia/Vant", "Vue3 中文文档", "https://cn.vuejs.org/guide/introduction.html", "Vant4 | Vue Router | Pinia", "h5-demo", "h5-demo", "在项目中定位 router/store/api 文件", "让 AI 加一页路由+一个封装 request", "☐ 五概念各能举例 ☐ 代码里能找到对应文件", 7),
            ("H5", "【目标5】浏览器/真机验收 + F12 排错", "H5 验收清单 v1", "Chrome 移动端模拟", "Chrome DevTools", "https://developer.chrome.com/docs/devtools/device-mode", "真机 USB 调试或扫码预览", "h5-demo", "h5-demo", "写验收项：布局/点击/loading/空态", "Chrome 手机模式+真机各测一遍；Console 排 1 个 bug", "☐ 验收清单可用 ☐ 会用 F12 定位问题", 7),
            ("H5", "【目标6+7】AI 代码审核 + AI 实现一页", "审核清单+AI 实现页", "AI 代码风险点", "Cursor 文档", "https://cursor.com/docs", "团队 AI Agent/Skill 流程", "h5-ai-page", "h5-ai-page", "写 AI 代码审核 10 条（XSS/密钥/硬编码等）", "提需求让 AI 做活动页→你按清单审核→改到能验收", "☐ 审核清单完成 ☐ AI 页浏览器验收通过", 7),
            ("H5", "【目标7+8】AI 工作流 + 后续计划 + W2 验收", "8 条目标自检表+后续计划", "E2E/部署概念", "Playwright 入门（了解）", "https://playwright.dev/docs/intro", "—", "h5-demo", "h5-demo", "走一遍：原型→文档→AI实现→验收→部署了解", "8 条目标逐条打勾；写个人后续 H5 学习计划", "☐ 8条全过 ☐ W2验收", 6),
        ],
    },
    "W3": {
        "阶段": "第1月·后端Java",
        "days": [
            ("Java", "JDK+Hello World", "程序能运行", "廖雪峰 Java 简介", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java | B站：Java 零基础", f"IDEA 新建 java-w3", "java-w3", "装 JDK17+IDEA", "Hello World", "☐ Run 成功", 7),
            ("Java", "变量与流程控制", "猜数字判断逻辑", "廖雪峰 流程控制", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java", "java-w3", "java-w3", "数据类型 if/else", "Scanner 输入", "☐ if 正确", 7),
            ("Java", "循环", "猜数字可玩", "廖雪峰 循环", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java", "java-w3", "java-w3", "for/while", "最多7次猜中", "☐ 游戏可玩", 7),
            ("Java", "类与对象", "Book 类", "廖雪峰 面向对象", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java", "java-w3", "java-w3", "封装", "控制台菜单", "☐ 图书类OK", 7),
            ("Java", "继承与接口", "图书借还系统", "廖雪峰 继承", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java", "java-w3", "java-w3", "继承/接口", "借还逻辑", "☐ 借还通", 7),
            ("Java", "集合", "学生 CRUD", "廖雪峰 集合", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java", "java-w3", "java-w3", "ArrayList/HashMap", "增删查", "☐ 学生系统OK", 7),
            ("Java", "异常+W3验收", "三程序可运行", "廖雪峰 异常", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "—", "java-w3", "java-w3", "try-catch", "push 代码", "☐ W3全过", 6),
        ],
    },
    "W4": {
        "阶段": "第1月·后端API",
        "days": [
            ("MySQL", "建表+ER", "shop 四表+ER图", "CREATE TABLE", "MySQL 文档", "https://dev.mysql.com/doc/", "SQL必知必会 | 菜鸟 SQL", f"Navicat 建库 sql-w4", "sql-w4", "CREATE TABLE", "测试数据", "☐ 表建好", 7),
            ("MySQL", "DML+查询", "10条 SQL 模板", "SQL DML", "MySQL 文档", "https://dev.mysql.com/doc/", "菜鸟 SQL", "sql-w4", "sql-w4", "INSERT/UPDATE/SELECT", "WHERE/GROUP BY", "☐ 10条模板", 7),
            ("MySQL", "JOIN+事务", "多表联查 demo", "SQL JOIN", "MySQL 文档", "https://dev.mysql.com/doc/", "菜鸟 SQL", "sql-w4", "sql-w4", "INNER/LEFT JOIN", "COMMIT/ROLLBACK", "☐ JOIN+事务", 7),
            ("Spring", "Spring Boot 创建", "项目启动成功", "Spring Boot 入门", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "start.spring.io | B站：Spring Boot3", f"start.spring.io → shop-api", "shop-api", "生成 shop-api", "application.yml 连库", "☐ 启动OK", 7),
            ("Spring", "MyBatis-Plus CRUD", "GET 列表 Postman 200", "MyBatis-Plus", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "baomidou.com", "shop-api", "shop-api", "Entity+Mapper", "GET /products", "☐ 列表有JSON", 7),
            ("Spring", "REST 四接口", "CRUD 全通", "RESTful", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "baomidou.com", "shop-api", "shop-api", "POST/PUT/DELETE", "Result 封装", "☐ CRUD全通", 7),
            ("Spring", "M1 后端验收", "API+SQL 文档", "—", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "—", "shop-api", "shop-api", "Postman Collection", "导出 shop.sql", "☐ M1达成", 6),
        ],
    },
    "W5": {
        "阶段": "第2月·Android",
        "days": [
            ("Android", "Android Studio+Hello", "模拟器跑通", "Android 开发者指南", "Android Docs", "https://developer.android.com/guide", "B站：Android 零基础 2024", f"New Project hello-android", "hello-android", "装 AS+SDK", "Hello World", "☐ 模拟器运行", 7),
            ("Android", "Activity 与 Layout", "两个页面跳转", "Activity 生命周期", "Android Docs", "https://developer.android.com/guide/components/activities", "菜鸟 Android", "hello-android", "hello-android", "LinearLayout/Constraint", "Intent 跳转", "☐ 跳转正常", 7),
            ("Android", "常用控件", "登录表单 UI", "TextView/Button", "Android Docs", "https://developer.android.com/guide/topics/ui", "B站：Android UI", "android-ui", "android-ui", "EditText/Button", "简单校验", "☐ 表单可输入", 7),
            ("Android", "RecyclerView 列表", "商品列表页", "RecyclerView", "Android Docs", "https://developer.android.com/develop/ui/views/layout/recyclerview", "B站：RecyclerView", "android-shop", "android-shop", "Adapter+ViewHolder", "mock 数据", "☐ 列表滚动", 7),
            ("Android", "Retrofit 网络", "调 shop-api 列表", "Retrofit", "Retrofit", "https://square.github.io/retrofit/", "B站：Retrofit 入门", "android-shop", "android-shop", "依赖+接口定义", "Gson 解析", "☐ API 有数据", 7),
            ("Android", "详情页+加购", "列表→详情通", "—", "Android Docs", "https://developer.android.com/guide", "—", "android-shop", "android-shop", "详情 Activity", "传递 id 参数", "☐ 详情可开", 7),
            ("Android", "W5 验收", "Android 调通 API", "—", "Android Docs", "https://developer.android.com/guide", "—", "android-shop", "android-shop", "自测主流程", "打 debug APK", "☐ W5全过", 6),
        ],
    },
    "W6": {
        "阶段": "第2月·Android进阶",
        "days": [
            ("Android", "Fragment+BottomNav", "三 Tab 结构", "Fragment", "Android Docs", "https://developer.android.com/guide/fragments", "B站：Fragment 导航", "android-shop", "android-shop", "BottomNavigation", "三 Fragment", "☐ Tab 可切", 7),
            ("Android", "SharedPreferences", "登录态本地保存", "数据存储", "Android Docs", "https://developer.android.com/training/data-storage", "B站：Android 存储", "android-shop", "android-shop", "存 token", "启动页判断", "☐ 重启仍登录", 7),
            ("Android", "Glide 图片", "列表带图", "Glide", "Glide", "https://bumptech.github.io/glide/", "B站：Glide 使用", "android-shop", "android-shop", "加载网络图", "占位/错误图", "☐ 图片显示", 7),
            ("Android", "权限与打包", "debug APK 安装", "应用权限", "Android Docs", "https://developer.android.com/guide/topics/permissions", "B站：Android 打包", "android-shop", "android-shop", "网络权限", "Build APK", "☐ 真机可装", 7),
            ("Android", "JWT 登录对接", "登录拿 token", "OkHttp 拦截器", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "B站：Android JWT", "android-shop", "android-shop", "/login 接口", "Header 带 token", "☐ 登录OK", 7),
            ("Android", "购物车本地", "加购数量合计", "—", "Android Docs", "https://developer.android.com/guide", "—", "android-shop", "android-shop", "Room/Simple 缓存", "合计金额", "☐ 购物车可用", 7),
            ("Android", "M2-Android 验收", "完整购物流程", "—", "Android Docs", "https://developer.android.com/guide", "—", "android-shop", "android-shop", "录屏演示", "push 代码", "☐ M2-Android", 6),
        ],
    },
    "W7": {
        "阶段": "第2月·Cocos",
        "days": [
            ("Cocos", "安装 Creator+界面", "空场景能预览", "Cocos 快速上手", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "B站：Cocos Creator 3 入门", f"新建 cocos-click-game", "cocos-click-game", "装 Creator 3.x", "场景+相机", "☐ 预览正常", 7),
            ("Cocos", "节点与组件", "Sprite 显示图片", "节点系统", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "B站：Cocos 节点", "cocos-click-game", "cocos-click-game", "Node/Component", "Sprite 换图", "☐ 精灵可见", 7),
            ("Cocos", "TypeScript 脚本", "点击改分数", "脚本基础", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "TS 快速入门", "cocos-click-game", "cocos-click-game", "@property 绑定", "onLoad/start", "☐ 点击加分", 7),
            ("Cocos", "UI 与 Label", "计分板+按钮", "UI 系统", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "B站：Cocos UI", "cocos-click-game", "cocos-click-game", "Canvas+Label", "Button 事件", "☐ UI 更新", 7),
            ("Cocos", "动画与 tween", "得分动画反馈", "动画系统", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "B站：Cocos tween", "cocos-click-game", "cocos-click-game", "Animation/tween", "缩放反馈", "☐ 有动画", 7),
            ("Cocos", "碰撞/定时生成", "点击得分小游戏", "物理碰撞", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "B站：Cocos 小游戏", "cocos-click-game", "cocos-click-game", "Collider", "倒计时结束", "☐ 小游戏可玩", 7),
            ("Cocos", "W7 验收", "Web 发布可玩", "构建发布", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "—", "cocos-click-game", "cocos-click-game", "构建 Web", "录屏 30 秒", "☐ W7全过", 6),
        ],
    },
    "W8": {
        "阶段": "第2月·Unity3D",
        "days": [
            ("Unity", "安装 Hub+创建项目", "3D 场景能跑", "Unity 入门", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "B站：Unity 2022 入门", f"Hub 新建 unity-roll-a-ball", "unity-roll-a-ball", "装 Unity Hub+LTS", "Sample 场景", "☐ Play 正常", 7),
            ("Unity", "GameObject 与 Transform", "移动立方体", "Transform", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "B站：Unity 基础", "unity-roll-a-ball", "unity-roll-a-ball", "Position/Rotation", "Scene 视图操作", "☐ 物体可移", 7),
            ("Unity", "C# 脚本基础", "键盘控制移动", "MonoBehaviour", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "菜鸟 C# | B站：Unity C#", "unity-roll-a-ball", "unity-roll-a-ball", "Update+Input", "刚体移动", "☐ WASD 可动", 7),
            ("Unity", "碰撞与触发", "收集金币得分", "Collider/Rigidbody", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "B站：Unity 碰撞", "unity-roll-a-ball", "unity-roll-a-ball", "OnTriggerEnter", "计分 UI", "☐ 碰撞加分", 7),
            ("Unity", "UI Canvas", "开始/结束界面", "Unity UI", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "B站：Unity UGUI", "unity-roll-a-ball", "unity-roll-a-ball", "Canvas+Text", "Button 重开", "☐ UI 可点", 7),
            ("Unity", "预制体 Prefab", "批量生成障碍", "Prefab", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "B站：Unity Prefab", "unity-roll-a-ball", "unity-roll-a-ball", "做 Prefab", "Instantiate", "☐ 障碍生成", 7),
            ("Unity", "M2-Unity 验收", "滚球小游戏可玩", "—", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "—", "unity-roll-a-ball", "unity-roll-a-ball", "打包 PC/WebGL 了解", "录屏+push", "☐ M2-Unity", 6),
        ],
    },
    "W9": {
        "阶段": "第3月·后端进阶",
        "days": [
            ("Spring", "JWT 注册登录", "token 链路", "JWT 入门", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "B站：Spring Boot JWT", "shop-api", "shop-api", "/register /login", "拦截器 401", "☐ 登录拿token", 7),
            ("Redis", "安装+五类型", "redis-cli 会用", "Redis 命令", "Redis", "https://redis.io/docs/", "菜鸟 Redis", "shop-api", "shop-api", "装 Redis", "String/Hash/ZSet", "☐ 类型练过", 7),
            ("Redis", "Spring 缓存", "商品列表缓存", "Spring Data Redis", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "B站：Redis 缓存", "shop-api", "shop-api", "RedisTemplate", "缓存接口", "☐ 缓存生效", 7),
            ("Spring", "订单/状态业务", "下单接口", "—", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "—", "shop-api", "shop-api", "订单表", "状态流转", "☐ 下单通", 7),
            ("Spring", "文件上传", "头像/商品图", "Multipart", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "B站：Spring 文件上传", "shop-api", "shop-api", "上传接口", "静态资源映射", "☐ 上传可访问", 7),
            ("Spring", "API 文档完善", "Apifox Collection", "Postman", "Apifox", "https://apifox.com/help/", "Postman 文档", "shop-api", "shop-api", "整理接口", "环境变量", "☐ 文档完成", 7),
            ("后端", "W9 回归测试", "Postman 全绿", "—", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "—", "shop-api", "shop-api", "Collection 跑", "修 bug", "☐ W9全过", 6),
        ],
    },
    "W10": {
        "阶段": "第3月·多端联调",
        "days": [
            ("设计", "多端需求+API 清单", "api-design.md", "—", "自建文档", "api-design.md", "B站：前后端分离设计", f"新建 game-platform 目录", "game-platform", "功能清单", "15 个接口", "☐ ER+API", 7),
            ("H5", "【联调】真实 API + 浏览器验收", "登录+列表 Postman/浏览器双验", "CORS", "MDN CORS", "https://developer.mozilla.org/zh-CN/docs/Web/HTTP/CORS", "Apifox 联调", "h5-demo", "h5-demo", "让 AI 改 baseURL/token 拦截器", "F12 Network 验收登录+列表；记录联调问题", "☐ H5 读写通 ☐ 验收清单全绿", 7),
            ("Android", "Android 联调修复", "主流程无阻塞", "—", "Android Docs", "https://developer.android.com/guide", "—", "android-shop", "android-shop", "对接新接口", "修 401/CORS", "☐ Android 主流程", 7),
            ("联调", "CORS/跨域排错", "联调记录表", "MDN CORS", "MDN CORS", "https://developer.mozilla.org/zh-CN/docs/Web/HTTP/CORS", "—", "—", "—", "前后端对齐", "修 P0 bug", "☐ 三端能登录", 7),
            ("联调", "购物车/下单闭环", "H5+Android 下单", "—", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "—", "game-platform", "game-platform", "H5 下单", "Android 下单", "☐ 下单闭环", 7),
            ("联调", "全流程 5 遍", "无 P0 缺陷", "—", "—", "—", "—", "—", "—", "自测脚本", "演示稿", "☐ W10全过", 7),
            ("联调", "UI/体验打磨", "loading/空态/错误", "—", "Vant4", "https://vant-ui.github.io/vant/#/zh-CN", "—", "h5-demo", "h5-demo", "按验收清单补空态/错误提示", "Android 提示统一", "☐ 体验可", 6),
        ],
    },
    "W11": {
        "阶段": "第3月·运维认知",
        "days": [
            ("运维", "【目标1】运维岗位分工与协作边界", "岗位协作一页纸", "运维/DevOps/SRE/DBA 分工", "运维岗位概览", "https://developer.aliyun.com/article/", "B站：运维是做什么的", f"新建 {CODE_ROOT}\\ops-notes", "ops-notes", "梳理：谁管机器/网络/发布/数据库/监控", "画协作图：开发提需求→运维怎么配合", "☐ 能讲清 4 类岗位 ☐ 协作图完成", 7),
            ("运维", "【目标2·基础】基础设施+网络概念", "基础概念对照表（20词）", "VPC/负载均衡/DNS/端口", "阿里云/腾讯云文档", "https://help.aliyun.com/", "B站：网络基础 运维必知", "ops-notes", "ops-notes", "学：服务器/虚拟机/容器/域名/HTTPS/防火墙", "对照 shop-api 部署图画：用户→Nginx→API→DB", "☐ 20 词能解释 ☐ 能画请求路径", 7),
            ("运维", "【目标2·安全】安全与权限概念", "安全概念清单", "安全组/ACL/密钥/HTTPS", "OWASP 入门", "https://owasp.org/www-project-top-ten/", "—", "ops-notes", "ops-notes", "学：安全组/白名单/密钥/HTTPS/最小权限", "写：shop-api 上线前 5 条安全检查", "☐ 安全清单完成 ☐ 知道密钥不能进 Git", 7),
            ("运维", "【目标2·发布+库】发布流程与数据库运维概念", "发布流程图+DB 概念表", "CI/CD/回滚/备份", "Docker 入门", "https://docs.docker.com/get-started/", "B站：Docker 概念 15min", f"docker compose 起 MySQL+Redis（验证概念）", "ops-docker", "学：蓝绿/滚动/回滚/mysqldump/主从（概念）", "写 shop-api 从 build→部署→回滚 流程", "☐ 发布流程能口述 ☐ compose 起库成功", 7),
            ("运维", "【目标3】故障沟通专业词汇", "运维术语表 30 条", "日志/告警/SLA/QPS", "SRE 词汇", "https://sre.google/sre-book/table-of-contents/", "—", "ops-glossary", "ops-glossary", "整理 30 条：502/504/OOM/磁盘/连接池/超时…", "模拟 3 段故障群聊：读懂并回复该问什么", "☐ 术语表 30 条 ☐ 群聊模拟完成", 7),
            ("运维", "【目标4】常见业务故障初判", "故障定位手册 v1", "排障分层思路", "Nginx 错误码", "https://nginx.org/en/docs/", "B站：线上故障排查思路", "ops-runbook", "ops-runbook", "写 6 类：502/504/接口慢/DB连不上/磁盘满/Redis挂", "每类写：现象→先查啥→找谁→临时措施", "☐ 6 类手册完成 ☐ 能口述排查顺序", 7),
            ("运维", "【目标5】自查题+后续计划+W11验收", "5条目标自检+薄弱点清单", "—", "自建 ops-quiz.md", "ops-quiz.md", "—", "ops-notes", "ops-notes", "50 道自查选择题/问答题", "5 条目标逐条打勾；写后续 3 个月运维深入计划", "☐ 5条全过 ☐ 薄弱点清单 ☐ W11验收", 6),
        ],
    },
    "W12": {
        "阶段": "第3月·结业",
        "days": [
            ("测试", "Postman 全量回归", "Collection 绿", "Postman", "Postman", "https://learning.postman.com/docs/introduction/overview/", "—", "—", "—", "Runner", "修接口", "☐ API 全绿", 7),
            ("测试", "Android+H5 冒烟", "主流程清单", "—", "—", "—", "—", "—", "—", "H5 冒烟", "Android 冒烟", "☐ 冒烟通过", 7),
            ("作品集", "GitHub README", "七方向架构图", "—", "—", "—", "—", "—", "—", "各项目 README", "无密钥泄露", "☐ README 完整", 7),
            ("演示", "3 分钟录屏", "多端 demo 视频", "—", "—", "—", "—", "—", "—", "H5+Android+游戏", "后端 API 展示", "☐ 有录屏", 7),
            ("复盘", "三个月总结", "总结文档", "—", "—", "—", "—", "—", "—", "薄弱项", "后续方向", "☐ 总结完成", 7),
            ("答辩", "模拟答辩", "能讲清七方向", "—", "—", "—", "—", "—", "—", "录音练习", "技术问答", "☐ 能3分钟介绍", 7),
            ("结业", "终极验收 M3", "全方向交付", "—", "—", "—", "—", "—", "—", "完整演示", "代码备份", "☐ 结业", 6),
        ],
    },
}

WEEK_GOALS = [
    ("W1", "工具+F12+读码（不对抗系统学HTML）", "读码笔记+Git+MySQL", "Chrome DevTools", "https://developer.chrome.com/docs/devtools", "MDN 只读查阅", "—", "为 H5 验收打基础"),
    ("W2", "H5 八条目标（AI实现·你来验收）", "8条自检表+验收清单+AI页", "Vue3 + Vant4", "https://cn.vuejs.org/", "Cursor/团队 AI Skill", "—", "见 sheet「H5学习目标」"),
    ("W3", "Java 基础", "猜数字+图书+学生", "廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "菜鸟 Java", "B站：Java 零基础", "后端Java"),
    ("W4", "MySQL+Spring Boot CRUD", "shop-api+SQL", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "MyBatis-Plus", "B站：Spring Boot3", "后端API"),
    ("W5", "Android 入门", "列表+调API", "Android Docs", "https://developer.android.com/guide", "Retrofit", "B站：Android 零基础", "Android"),
    ("W6", "Android 进阶", "登录+购物车 APK", "Android Docs", "https://developer.android.com/guide", "Glide/JWT", "B站：Android 项目", "Android"),
    ("W7", "Cocos Creator", "点击得分小游戏", "Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "TypeScript", "B站：Cocos Creator3", "Cocos"),
    ("W8", "Unity3D", "滚球收集游戏", "Unity Manual", "https://docs.unity3d.com/Manual/index.html", "C# 基础", "B站：Unity 入门", "Unity"),
    ("W9", "JWT+Redis+业务API", "完整 shop-api", "Spring Boot", "https://docs.spring.io/spring-boot/reference/", "Redis", "B站：Spring Redis JWT", "后端进阶"),
    ("W10", "H5+Android+后端联调", "三端主流程通", "MDN CORS", "https://developer.mozilla.org/zh-CN/docs/Web/HTTP/CORS", "api-design.md", "B站：前后端联调", "多端联调"),
    ("W11", "运维五条目标（认知+协作+初判）", "5条自检+术语表+故障手册", "云厂商文档+Docker概念", "https://docs.docker.com/get-started/", "SRE/Google", "—", "见 sheet「运维学习目标」"),
    ("W12", "测试+作品集+结业", "录屏+README+答辩", "Postman", "https://learning.postman.com/docs/introduction/overview/", "GitHub Actions", "B站：项目部署", "结业"),
]

OVERVIEW = [
    ("第1月", "H5 验收向 + Java/Spring 入门", "H5 八条达标 + shop-api", "F12 读码 AI验收 Vue Router Pinia Java Spring"),
    ("第2月", "Android + Cocos + Unity 三端", "Android APK + 2个小游戏", "Android Retrofit Cocos Unity C#"),
    ("第3月", "后端进阶 + 联调 + 运维认知", "三端联调 + 运维五目标达标 + 作品集", "故障定位 术语 协作边界 Docker概念"),
]

MILESTONES = [
    ("M1", "W4结束", "H5八条+后端API", "8条自检全过+shop-api Postman通"),
    ("M2", "W8结束", "Android+游戏引擎", "Android APK+Cocos+Unity 各1个"),
    ("M3", "W12结束", "联调+运维认知+结业", "三端通+运维5条达标+README+录屏"),
]

TOOLS = [
    ("Chrome 浏览器", "F12 验收排错", "https://www.google.com/chrome/", "D1", "必装"),
    ("VS Code / Cursor", "编辑与 AI", "https://cursor.com", "D1", "必装"),
    ("Git", "版本管理", "https://git-scm.com", "D1", "必装"),
    ("Postman/Apifox", "测接口", "https://www.postman.com/downloads", "D1", "必装"),
    ("MySQL 8", "数据库", "https://dev.mysql.com/downloads", "D3", "W1"),
    ("Node.js LTS", "H5/Vant", "https://nodejs.org", "W2", "H5前"),
    ("JDK 17", "Java", "https://adoptium.net", "W3", "Java前"),
    ("IDEA Community", "Java/Spring", "https://www.jetbrains.com/idea/download", "W3", "Java前"),
    ("Android Studio", "Android", "https://developer.android.com/studio", "W5", "Android前"),
    ("Cocos Creator 3.x", "2D游戏", "https://www.cocos.com/creator-download", "W7", "Cocos前"),
    ("Unity Hub + LTS", "3D游戏", "https://unity.com/download", "W8", "Unity前"),
    ("Redis", "缓存", "https://redis.io/download", "W9", "后端进阶"),
    ("Docker Desktop", "验证容器概念", "https://www.docker.com/products/docker-desktop", "W11", "可选动手"),
]

LINKS = [
    ("Chrome DevTools", "https://developer.chrome.com/docs/devtools", "W1-W2,W10"),
    ("Vue3 中文", "https://cn.vuejs.org/", "W2,W10"),
    ("Vant4", "https://vant-ui.github.io/vant/#/zh-CN", "W2,W10"),
    ("Cursor 文档", "https://cursor.com/docs", "W2"),
    ("廖雪峰 Java", "https://liaoxuefeng.com/books/java/index.html", "W3"),
    ("Spring Boot", "https://docs.spring.io/spring-boot/reference/", "W4,W9"),
    ("Android 开发者", "https://developer.android.com/guide", "W5-W6,W10"),
    ("Cocos 文档", "https://docs.cocos.com/creator/manual/zh/", "W7"),
    ("Unity Manual", "https://docs.unity3d.com/Manual/index.html", "W8"),
    ("Docker 入门", "https://docs.docker.com/get-started/", "W11"),
    ("Nginx", "https://nginx.org/en/docs/", "W11-W12"),
    ("GitHub Actions", "https://docs.github.com/actions", "W11"),
]

BOOKS = [
    ("《SQL必知必会》", "书", "W4", "练查询"),
    ("Chrome F12 移动端调试", "文档", "W1-W2", "验收必备"),
    ("Cursor AI 辅助 H5", "工作流", "W2", "实现交给AI"),
    ("Android 零基础", "B站关键词", "W5-W6", "跟做一套"),
    ("Cocos Creator3 入门", "B站关键词", "W7", "2D小游戏"),
    ("Unity 2022 入门", "B站关键词", "W8", "滚球游戏"),
    ("Linux Docker 运维", "B站关键词", "W11", "概念为主"),
    ("线上故障排查", "B站关键词", "W11", "定位思路"),
]

TRACK_PROGRESS_SPECS = [
    ("工具与读码", "W1", ["W1"], "F12读码+Git+MySQL"),
    ("H5", "W2", ["W2"], "八条目标自检全过"),
    ("后端", "W3+W4+W9", ["W3", "W4", "W9"], "shop-api+JWT/Redis"),
    ("Android", "W5+W6", ["W5", "W6"], "购物APK"),
    ("Cocos", "W7", ["W7"], "2D小游戏"),
    ("Unity3D", "W8", ["W8"], "滚球游戏"),
    ("多端联调", "W10", ["W10"], "H5+Android+后端"),
    ("运维", "W11", ["W11"], "五条目标自检全过"),
    ("结业", "W12", ["W12"], "录屏+README+答辩"),
]


def merge_day_group(group):
    """将同一天内的多个原日任务合并为一行（学时取末项，验收日仍为 6h）。"""
    if len(group) == 1:
        return group[0]
    mods = "·".join(dict.fromkeys(t[0] for t in group))
    goal = "；".join(t[1] for t in group)
    deliver = "；".join(t[2] for t in group)
    read = " / ".join(t[3] for t in group if t[3] and t[3] != "—") or "—"
    doc_name = group[-1][4]
    doc_link = group[-1][5]
    tut_parts = [t[6] for t in group if t[6] and t[6] != "—"]
    tut = " | ".join(dict.fromkeys(tut_parts)) if tut_parts else "—"
    create = next((t[7] for t in reversed(group) if t[7] and t[7] != "—"), "—")
    folder = next((t[8] for t in reversed(group) if t[8] and t[8] != "—"), "—")
    am = " → ".join(t[9] for t in group)
    pm = " → ".join(t[10] for t in group)
    check = " ".join(t[11] for t in group)
    hours = group[-1][12]
    return (mods, goal, deliver, read, doc_name, doc_link, tut, create, folder, am, pm, check, hours)


def compressed_week_days(wk):
    originals = WEEKS[wk]["days"]
    return [merge_day_group([originals[i] for i in g]) for g in WEEK_MERGE_GROUPS[wk]]


def milestone_day_ids(days):
    ids = {}
    for ms_id, wk in MILESTONE_WEEKS.items():
        wk_days = [d for d in days if d["周次"] == wk]
        ids[ms_id] = wk_days[-1]["序号"]
    return ids


def build_time_progress(days):
    start = days[0]["日期"]
    end = days[-1]["日期"]
    total_h = sum(int(d["建议学时"]) for d in days)
    calendar_weeks = round((date.fromisoformat(end) - date.fromisoformat(start)).days / 7 + 1)
    ms_labels = {
        "M1": "H5八条达标 + 后端API入门",
        "M2": "Android+Cocos+Unity各1项目",
        "M3": "联调+运维5条达标+结业",
    }
    rows = [
        ["总周期", f"{start} ~ {end}", f"{len(days)} 个工作日", "12 内容周", f"{total_h} 小时", f"约 {calendar_weeks} 个自然周"],
        ["学习日（周一至周五）", "7 小时/天", "上午 2.5h", "下午 2.5h", "晚上 0.5h", "弹性 1.5h"],
        ["周验收日（每内容周最后1个工作日）", "6 小时/天", "上午 2.5h", "下午 2.5h", "晚上 0.5h", "弹性 0.5h"],
        ["周末", "休息", "不排课", "", "", ""],
        ["说明", "原 84 天精简为 66 天", "合并同类任务", "交付物不变", "", ""],
        ["", "", "", "", "", ""],
        ["方向", "周次", "起止日期", "天数", "计划学时", "验收日 / 交付物"],
    ]
    for name, weeks_label, week_keys, deliver in TRACK_PROGRESS_SPECS:
        week_days = [d for d in days if d["周次"] in week_keys]
        dr = f"{week_days[0]['日期']} ~ {week_days[-1]['日期']}"
        check = week_days[-1]["日期"][5:]
        week_h = sum(int(d["建议学时"]) for d in week_days)
        rows.append([name, weeks_label, dr, str(len(week_days)), str(week_h), f"{check} · {deliver}"])
    rows.extend([
        ["", "", "", "", "", ""],
        ["里程碑", "日期", "对应天", "说明", "", ""],
    ])
    for ms_id, wk in MILESTONE_WEEKS.items():
        last = next(d for d in reversed(days) if d["周次"] == wk)
        rows.append([ms_id, last["日期"], last["序号"], ms_labels[ms_id], "", ""])
    return rows

USAGE_ROWS = [
    ["【七方向入门 · 完整学习手册】后端 / Android / H5 / Cocos / Unity / 运维", ""],
    ["", ""],
    ["你每天只做 3 步", ""],
    ["1", "打开 sheet「每日学习计划」→ 找到今天日期那一行"],
    ["2", "上午：看「今日阅读章节」+「主文档链接」15~30分钟 → 按「上午任务」执行"],
    ["3", "下午：做「交付物」→ 晚上对照「验收清单」→「完成」选「是」"],
    ["", ""],
    ["H5 / 运维 学习理念（重要）", ""],
    ["H5", H5_PHILOSOPHY],
    ["运维", OPS_PHILOSOPHY],
    ["", ""],
    ["其它 sheet 什么时候看", ""],
    ["H5学习目标", "W2 开始前通读；每天对照当日目标编号"],
    ["运维学习目标", "W11 开始前通读；每天对照当日目标编号"],
    ["学习时间与进度", "查看总学时、各方向进度、里程碑"],
    ["怎么学", "第一次学习前读；H5/运维周对照专属循环"],
    ["周目标与参考书", "每周一开始看本周一行"],
    ["工具安装", "D1/D3/W2/W3/W5/W7/W8/W11 等安装日对照"],
    ["里程碑", "W4/W8/W12 验收日对照 M1/M2/M3"],
    ["", ""],
    ["代码根目录", CODE_ROOT],
    ["准确时间", "66 个工作日；6/1 起仅周一至周五，8/31 前结业"],
    ["三个月目标", "H5 八条 + 后端 API + Android + Cocos + Unity + 运维五条 + 三端联调 + 作品集"],
]

HOW_TO_LEARN = [
    ["步骤", "动作", "具体做法"],
    ["通用", "看", "文档/教程 15~30 分钟，只读与今天相关的一节"],
    ["通用", "敲", "自己打字敲示例，不要只复制（后端/Android/游戏周）"],
    ["通用", "跑", "浏览器/Postman/模拟器/启动项目看结果"],
    ["通用", "改", "故意改一点代码观察变化"],
    ["H5专周", "循环", H5_LEARN_LOOP],
    ["运维专周", "循环", OPS_LEARN_LOOP],
    ["卡住", "顺序", "读报错最后一行 → 查官方文档 → 对比教程 → 求助AI"],
    ["", "", ""],
    ["每日时间建议", "", ""],
    ["上午 2.5h", "看资料 + 跟做/验收", ""],
    ["下午 2.5h", "做交付物", ""],
    ["晚上 0.5h", "验收打勾 + 3句复盘", ""],
]

QUICK_7_DAYS = [
    ["天", "只看这一节", "只做这一件", "文件夹"],
    ["1", "H5 八条目标 + Chrome F12 入门", "装工具；抄 8 条目标", "day01-intro"],
    ["2", "Pro Git 第1章", "git init + commit", "day02-git"],
    ["3", "MySQL 安装（了解即可）", "能连上数据库", "—"],
    ["4", "HTML 结构 + CSS 适配（只读）", "F12 标注结构；手机模式看适配", "h5-read-html"],
    ["5", "Chrome Network 面板", "跟一条请求链路", "h5-read-network"],
    ["6", "复盘 + 看 W2 H5专周", "预习 D7 项目目录", "—"],
]

H5_GOAL_DAYS = ["D7", "D8", "D9", "D10", "D11", "D11", "D12", "D12"]
OPS_GOAL_DAYS = ["D57", "D58", "D59", "D60", "D61"]

FAQ_ROWS = [
    ["问题", "回答"],
    ["怎么开始?", "打开「每日学习计划」今天一行；H5/运维周先看对应「学习目标」sheet"],
    ["H5 要先学完 HTML 吗?", "不必。先懂 H5 在干什么、写清需求，实现交给 AI，你在浏览器验收"],
    ["运维要先考证书吗?", "不必。先懂岗位边界、术语、故障初判；Docker 练一次验证概念即可"],
    ["教程和计划不符?", "以「交付物/验收清单」为准，教程多讲的跳过"],
    ["怎么创建项目?", "看「怎么创建」列；H5 用 npm create vite；后端用 start.spring.io"],
    ["英文看不懂?", "优先 MDN中文、Vue cn、廖雪峰；命令保持英文"],
    ["要学多久找工作?", "本计划=七方向入门+可演示项目；就业需专精一个方向+更多项目"],
]


def build_usage_rows(days):
    start = days[0]["日期"]
    end = days[-1]["日期"]
    rows = []
    for row in USAGE_ROWS:
        r = list(row)
        if r[0] == "准确时间":
            total_h = sum(int(d["建议学时"]) for d in days)
            r[1] = (
                f"{total_h} 小时 / {len(days)} 个工作日（{start} ~ {end}）"
                "；仅周一至周五，周末休息"
            )
        rows.append(r)
    return rows


def write_usage_sheet(ws, days):
    ws.delete_rows(1, ws.max_row)
    for row in build_usage_rows(days):
        ws.append(row)


def write_how_to_learn(ws):
    ws.delete_rows(1, ws.max_row)
    for row in HOW_TO_LEARN:
        ws.append(row)


def write_quick_7_days(ws):
    ws.delete_rows(1, ws.max_row)
    for row in QUICK_7_DAYS:
        ws.append(row)


def write_faq(ws):
    ws.delete_rows(1, ws.max_row)
    for row in FAQ_ROWS:
        ws.append(row)


def write_goal_sheet(ws, title_philosophy, goals, day_hints):
    ws.delete_rows(1, ws.max_row)
    ws.append(["序号", "学习目标", "对应周/天", "验收方式"])
    ws.append(["理念", title_philosophy, "全程", "能口述核心理念"])
    for idx, goal in enumerate(goals):
        day_hint = day_hints[idx] if idx < len(day_hints) else ""
        ws.append([idx + 1, goal, day_hint, "打勾+交付物"])


def build_days():
    days = []
    i = 0
    for wk in [f"W{n}" for n in range(1, 13)]:
        info = WEEKS[wk]
        for tpl in compressed_week_days(wk):
            d = learning_date(START, i)
            row = {
                "序号": f"D{i + 1}",
                "日期": d.strftime("%Y-%m-%d"),
                "星期": WEEKDAY_CN[d.weekday()],
                "周次": wk,
                "阶段": info["阶段"],
                "模块": tpl[0],
                "今日目标": tpl[1],
                "交付物（必达）": tpl[2],
                "怎么学（固定循环）": (
                    H5_LEARN_LOOP if wk in ("W1", "W2") or tpl[0] == "H5"
                    else OPS_LEARN_LOOP if wk == "W11"
                    else LEARN_LOOP
                ),
                "今日阅读章节": tpl[3],
                "主文档（名称）": tpl[4],
                "主文档链接": tpl[5],
                "入门教程/视频": tpl[6],
                "怎么创建": tpl[7],
                "代码文件夹": tpl[8],
                "上午任务": tpl[9],
                "下午任务": tpl[10],
                "晚上复盘提示": REVIEW,
                "验收清单": tpl[11],
                "建议学时": tpl[12],
                "完成": "",
                "实际学时": "",
                "备注": "",
            }
            days.append(row)
            i += 1
    last = date.fromisoformat(days[-1]["日期"])
    if len(days) != TOTAL_LEARNING_DAYS:
        raise ValueError(f"学习日应为 {TOTAL_LEARNING_DAYS} 天，实际 {len(days)} 天")
    if last > END_DEADLINE:
        raise ValueError(f"计划结束日 {last} 超过 {END_DEADLINE}，请调整 START 或总学习日数")
    return days


def week_date_ranges(days):
    ranges = {}
    for row in days:
        wk = row["周次"]
        ranges.setdefault(wk, []).append(row["日期"])
    return {k: (v[0], v[-1]) for k, v in ranges.items()}


def write_sheet_rows(ws, rows, clear_from=2):
    if ws.max_row >= clear_from:
        ws.delete_rows(clear_from, ws.max_row - clear_from + 1)
    for r, row in enumerate(rows, start=clear_from):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)


def main():
    days = build_days()
    ranges = week_date_ranges(days)
    wb = load_workbook(XLSX)

    headers = [
        "序号", "日期", "星期", "周次", "阶段", "模块", "今日目标", "交付物（必达）",
        "怎么学（固定循环）", "今日阅读章节", "主文档（名称）", "主文档链接", "入门教程/视频",
        "怎么创建", "代码文件夹", "上午任务", "下午任务", "晚上复盘提示", "验收清单",
        "建议学时", "完成", "实际学时", "备注",
    ]
    ws = wb["每日学习计划"]
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in days:
        ws.append([row.get(h, "") for h in headers])

    ws2 = wb["周目标与参考书"]
    week_rows = []
    for item in WEEK_GOALS:
        wk, goal, deliver, doc, link, tut, vid, note = item
        dr = ranges[wk]
        week_rows.append([wk, f"{dr[0]} ~ {dr[1]}", goal, deliver, doc, link, tut, vid, note])
    write_sheet_rows(ws2, week_rows)

    ws3 = wb["三个月总览"]
    write_sheet_rows(ws3, OVERVIEW)

    ws4 = wb["里程碑"]
    ms_rows = []
    for node, week, goal, verify in MILESTONES:
        last = next(d for d in reversed(days) if d["周次"] == MILESTONE_WEEKS[node])
        ms_rows.append([node, last["日期"], week, goal, verify, ""])
    write_sheet_rows(ws4, ms_rows)

    ws5 = wb["工具安装"]
    tool_rows = [[t[0], t[1], t[2], t[3], t[4], ""] for t in TOOLS]
    write_sheet_rows(ws5, tool_rows)

    ws6 = wb["链接速查"]
    write_sheet_rows(ws6, LINKS)

    ws7 = wb["书单与视频"]
    write_sheet_rows(ws7, BOOKS)

    if "使用说明" in wb.sheetnames:
        write_usage_sheet(wb["使用说明"], days)

    if "怎么学" in wb.sheetnames:
        write_how_to_learn(wb["怎么学"])

    if "7天极简入门" in wb.sheetnames:
        write_quick_7_days(wb["7天极简入门"])

    if "常见问题" in wb.sheetnames:
        write_faq(wb["常见问题"])

    if "运维学习目标" in wb.sheetnames:
        ws_ops = wb["运维学习目标"]
    else:
        ws_ops = wb.create_sheet("运维学习目标")
    write_goal_sheet(ws_ops, OPS_PHILOSOPHY, OPS_GOALS, OPS_GOAL_DAYS)

    if "H5学习目标" in wb.sheetnames:
        ws_h5 = wb["H5学习目标"]
    else:
        ws_h5 = wb.create_sheet("H5学习目标")
    write_goal_sheet(ws_h5, H5_PHILOSOPHY, H5_GOALS, H5_GOAL_DAYS)

    if "学习时间与进度" in wb.sheetnames:
        ws_t = wb["学习时间与进度"]
        ws_t.delete_rows(1, ws_t.max_row)
    else:
        ws_t = wb.create_sheet("学习时间与进度")
    for row in build_time_progress(days):
        ws_t.append(row)

    try:
        wb.save(XLSX)
        print(f"已保存: {XLSX}")
    except PermissionError:
        print(f"警告: Excel 被占用，请关闭「三个月全栈入门计划.xlsx」后重新运行 rebuild_multi_track_plan.py")

    from build_html import build_html, HTML_OUT
    html = build_html(days)
    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"已生成 HTML: {HTML_OUT}（共 {len(days)} 天）")
    end = date.fromisoformat(days[-1]["日期"])
    print(f"已重建 {len(days)} 个工作日: {days[0]['日期']} ~ {end.isoformat()}（周末休息）")


if __name__ == "__main__":
    main()
