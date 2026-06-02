# -*- coding: utf-8 -*-
"""从 Excel 生成交互式 HTML 学习计划（自动保存学习记录）"""

import json
import os
import re
from datetime import date
from openpyxl import load_workbook

DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DIR, "三个月全栈入门计划.xlsx")
HTML_OUT = os.path.join(DIR, "三个月全栈入门计划.html")
STORAGE_KEY = "stack-fullstack-plan-v2"

TRACK_SPECS = [
    {"id": "base", "name": "工具与读码", "weeks": ["W1"], "deliver": "F12 读码 + Git + MySQL（不系统写 HTML）"},
    {"id": "h5", "name": "H5", "weeks": ["W2"], "deliver": "八条学习目标自检全过 + 验收清单 + AI 实现页"},
    {"id": "backend", "name": "后端", "weeks": ["W3", "W4", "W9"], "deliver": "Java 三程序 + shop-api + JWT/Redis"},
    {"id": "android", "name": "Android", "weeks": ["W5", "W6"], "deliver": "购物 APK（列表/登录/购物车）"},
    {"id": "cocos", "name": "Cocos", "weeks": ["W7"], "deliver": "点击得分 2D 小游戏（Web 发布）"},
    {"id": "unity", "name": "Unity3D", "weeks": ["W8"], "deliver": "滚球收集 3D 小游戏"},
    {"id": "devops", "name": "运维", "weeks": ["W11"], "deliver": "5条目标自检全过 + 术语表 + 故障定位手册"},
    {"id": "integration", "name": "多端联调", "weeks": ["W10"], "deliver": "H5 + Android + 后端三端主流程"},
    {"id": "graduate", "name": "结业", "weeks": ["W12"], "deliver": "测试 + README + 录屏 + 答辩"},
]

MILESTONE_WEEKS = [
    ("M1", "W4", "H5 八条达标 + 后端 API 入门"),
    ("M2", "W8", "Android + Cocos + Unity 各 1 项目"),
    ("M3", "W12", "三端联调 + 运维五目标达标 + 结业"),
]


def compute_plan_meta(days):
    start = days[0]["日期"]
    end = days[-1]["日期"]
    total_h = sum(int(d["建议学时"]) for d in days)
    calendar_weeks = round((date.fromisoformat(end) - date.fromisoformat(start)).days / 7 + 1)
    milestones = []
    for ms_id, wk, label in MILESTONE_WEEKS:
        last = next(d for d in reversed(days) if d["周次"] == wk)
        milestones.append({
            "id": ms_id,
            "date": last["日期"],
            "dayId": last["序号"],
            "label": label,
        })
    return {
        "totalDays": len(days),
        "totalHours": total_h,
        "startDate": start,
        "endDate": end,
        "weeksCount": 12,
        "calendarWeeks": calendar_weeks,
        "weeklyHours": round(total_h / 12),
        "weekdayHours": 7,
        "reviewDayHours": 6,
        "weekdayOnly": True,
        "dailySchedule": [
            ["上午", "2.5 小时", "看「今日阅读」+ 跟敲示例"],
            ["下午", "2.5 小时", "完成当日交付物"],
            ["晚上", "0.5 小时", "验收清单 + 三句复盘"],
            ["弹性", "1.5 小时", "补练习 / 卡顿消化（计入 7h）"],
        ],
        "milestones": milestones,
    }


def compute_track_defs(days):
    day_map = {d["序号"]: d["日期"] for d in days}
    tracks = []
    for spec in TRACK_SPECS:
        track_days = [d for d in days if d["周次"] in spec["weeks"]]
        ids = [d["序号"] for d in track_days]
        weeks_label = " + ".join(spec["weeks"]) if len(spec["weeks"]) > 1 else spec["weeks"][0]
        tracks.append({
            "id": spec["id"],
            "name": spec["name"],
            "hours": sum(int(d["建议学时"]) for d in track_days),
            "weeks": weeks_label,
            "dayIds": ids,
            "deliver": spec["deliver"],
            "dateStart": day_map[ids[0]],
            "dateEnd": day_map[ids[-1]],
            "checkDate": day_map[ids[-1]],
        })
    return tracks


def load_days():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["每日学习计划"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    days = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        if not re.match(r"^D\d+$", str(row[0]).strip()):
            continue
        d = {h: (v if v is not None else "") for h, v in zip(headers, row)}
        d["id"] = str(d.get("序号", ""))
        checklist = str(d.get("验收清单", ""))
        d["checklistItems"] = [
            x.strip().lstrip("☐").lstrip("☑").strip()
            for x in re.split(r"[;；\n]", checklist)
            if x.strip() and "☐" in x or len(x.strip()) > 2
        ]
        if not d["checklistItems"] and checklist:
            d["checklistItems"] = [checklist]
        days.append(d)
    wb.close()
    return days


def build_html(days):
    plan_meta = compute_plan_meta(days)
    track_defs = compute_track_defs(days)
    data_json = json.dumps(days, ensure_ascii=False)
    meta_json = json.dumps(plan_meta, ensure_ascii=False)
    tracks_json = json.dumps(track_defs, ensure_ascii=False)
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>三个月七方向入门计划</title>
  <style>
    :root {{
      --bg: #0f1419;
      --panel: #1a2332;
      --panel2: #243044;
      --text: #e7ecf3;
      --muted: #8b9cb3;
      --accent: #3dd6c3;
      --accent2: #5b8def;
      --done: #6bcb77;
      --warn: #ffb347;
      --border: #2d3a4f;
      --sidebar-w: 300px;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: var(--bg);
      color: var(--text);
      height: 100vh;
      overflow: hidden;
    }}
    .app {{ display: flex; height: 100vh; }}
    .sidebar {{
      width: var(--sidebar-w);
      min-width: var(--sidebar-w);
      background: var(--panel);
      border-right: 1px solid var(--border);
      display: flex;
      flex-direction: column;
    }}
    .sidebar-head {{
      padding: 16px;
      border-bottom: 1px solid var(--border);
    }}
    .sidebar-head h1 {{ font-size: 1rem; margin: 0 0 8px; line-height: 1.4; }}
    .progress-wrap {{ margin-top: 10px; }}
    .progress-bar {{
      height: 8px;
      background: var(--panel2);
      border-radius: 4px;
      overflow: hidden;
    }}
    .progress-fill {{
      height: 100%;
      background: linear-gradient(90deg, var(--accent), var(--accent2));
      width: 0%;
      transition: width .3s;
    }}
  .progress-text {{ font-size: 12px; color: var(--muted); margin-top: 6px; }}
    .toolbar {{ display: flex; flex-wrap: wrap; gap: 6px; margin-top: 10px; }}
    .btn {{
      border: 1px solid var(--border);
      background: var(--panel2);
      color: var(--text);
      padding: 6px 10px;
      border-radius: 6px;
      font-size: 12px;
      cursor: pointer;
    }}
    .btn:hover {{ border-color: var(--accent); }}
    .btn-primary {{ background: var(--accent); color: #0f1419; border-color: var(--accent); font-weight: 600; }}
    .search {{
      width: 100%;
      margin-top: 10px;
      padding: 8px 10px;
      border-radius: 6px;
      border: 1px solid var(--border);
      background: var(--bg);
      color: var(--text);
    }}
    .day-list {{
      flex: 1;
      overflow-y: auto;
      padding: 8px;
    }}
    .week-label {{
      font-size: 11px;
      color: var(--accent);
      padding: 10px 8px 4px;
      font-weight: 600;
      letter-spacing: .05em;
    }}
    .day-item {{
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 8px 10px;
      border-radius: 8px;
      cursor: pointer;
      font-size: 13px;
      border: 1px solid transparent;
    }}
    .day-item:hover {{ background: var(--panel2); }}
    .day-item.active {{
      background: rgba(61, 214, 195, .15);
      border-color: var(--accent);
    }}
    .day-item.done .day-id {{ color: var(--done); }}
    .day-dot {{
      width: 8px;
      height: 8px;
      border-radius: 50%;
      background: var(--border);
      flex-shrink: 0;
    }}
    .day-item.done .day-dot {{ background: var(--done); }}
    .day-item.partial .day-dot {{ background: var(--warn); }}
    .day-meta {{ color: var(--muted); font-size: 11px; }}
    .main {{
      flex: 1;
      overflow-y: auto;
      padding: 24px 32px 48px;
    }}
    .save-status {{
      position: fixed;
      top: 12px;
      right: 16px;
      font-size: 12px;
      color: var(--muted);
      z-index: 10;
    }}
    .save-status.saved {{ color: var(--done); }}
    .tabs {{ display: flex; gap: 8px; margin-bottom: 20px; }}
    .tab {{
      padding: 8px 16px;
      border-radius: 8px;
      cursor: pointer;
      background: var(--panel);
      border: 1px solid var(--border);
      font-size: 14px;
    }}
    .tab.active {{ background: var(--accent); color: #0f1419; border-color: var(--accent); font-weight: 600; }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 16px 20px;
      margin-bottom: 16px;
    }}
    .card h2 {{ margin: 0 0 12px; font-size: 1.1rem; color: var(--accent); }}
    .card h3 {{ margin: 0 0 8px; font-size: .95rem; }}
    .hero-title {{ font-size: 1.5rem; margin: 0 0 4px; }}
    .hero-sub {{ color: var(--muted); margin-bottom: 20px; }}
    .tags {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; }}
    .tag {{
      font-size: 12px;
      padding: 4px 10px;
      border-radius: 20px;
      background: var(--panel2);
      color: var(--muted);
    }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
    @media (max-width: 900px) {{
      .grid-2 {{ grid-template-columns: 1fr; }}
      .app {{ flex-direction: column; }}
      .sidebar {{ width: 100%; min-width: 0; max-height: 40vh; }}
    }}
    .field-label {{ font-size: 12px; color: var(--muted); margin-bottom: 4px; }}
    .field-value {{ line-height: 1.6; white-space: pre-wrap; }}
    a {{ color: var(--accent2); }}
    .record-section label {{ display: block; font-size: 13px; color: var(--muted); margin: 12px 0 6px; }}
    select, input[type="number"], textarea {{
      width: 100%;
      padding: 10px 12px;
      border-radius: 8px;
      border: 1px solid var(--border);
      background: var(--bg);
      color: var(--text);
      font-family: inherit;
      font-size: 14px;
    }}
    textarea {{ min-height: 100px; resize: vertical; }}
    .checklist {{ list-style: none; padding: 0; margin: 0; }}
    .checklist li {{
      display: flex;
      align-items: flex-start;
      gap: 10px;
      padding: 8px 0;
      border-bottom: 1px solid var(--border);
    }}
    .checklist li:last-child {{ border-bottom: none; }}
    .checklist input {{ margin-top: 4px; width: 18px; height: 18px; accent-color: var(--accent); }}
    .log-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    .log-table th, .log-table td {{
      padding: 10px;
      text-align: left;
      border-bottom: 1px solid var(--border);
    }}
    .log-table th {{ color: var(--muted); font-weight: 500; }}
    .empty {{ color: var(--muted); text-align: center; padding: 40px; }}
    .hidden {{ display: none !important; }}
    .track-bar {{ height: 10px; background: var(--panel2); border-radius: 5px; overflow: hidden; margin: 8px 0; }}
    .track-fill {{ height: 100%; background: linear-gradient(90deg, var(--accent2), var(--accent)); width: 0%; transition: width .3s; }}
    .track-row {{ margin-bottom: 18px; padding-bottom: 14px; border-bottom: 1px solid var(--border); }}
    .track-row:last-child {{ border-bottom: none; }}
    .track-head {{ display: flex; justify-content: space-between; align-items: baseline; gap: 12px; flex-wrap: wrap; }}
    .track-name {{ font-weight: 600; }}
    .track-meta {{ font-size: 12px; color: var(--muted); }}
    .timeline {{ border-left: 2px solid var(--border); margin-left: 8px; padding-left: 16px; }}
    .timeline-item {{ margin-bottom: 16px; position: relative; }}
    .timeline-item::before {{
      content: ""; position: absolute; left: -21px; top: 6px;
      width: 10px; height: 10px; border-radius: 50%; background: var(--accent);
    }}
    .timeline-item.done::before {{ background: var(--done); }}
    .timeline-item.future::before {{ background: var(--border); }}
  </style>
</head>
<body>
  <div class="save-status" id="saveStatus">记录将自动保存到本机浏览器</div>
  <div class="app">
    <aside class="sidebar">
      <div class="sidebar-head">
        <h1>三个月七方向入门计划</h1>
        <div class="progress-wrap">
          <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
          <div class="progress-text" id="progressText">进度 0/84</div>
        </div>
        <div class="toolbar">
          <button class="btn btn-primary" id="btnToday">今天</button>
          <button class="btn" id="btnExport">导出记录</button>
          <button class="btn" id="btnImport">导入记录</button>
        </div>
        <input type="search" class="search" id="search" placeholder="搜索 D1、H5、Android、Cocos…" />
      </div>
      <div class="day-list" id="dayList"></div>
    </aside>
    <main class="main">
      <div class="tabs">
        <div class="tab active" data-tab="plan">今日学习</div>
        <div class="tab" data-tab="progress">时间与进度</div>
        <div class="tab" data-tab="log">学习记录总览</div>
        <div class="tab" data-tab="help">使用说明</div>
      </div>
      <div id="panelPlan"></div>
      <div id="panelProgress" class="hidden"></div>
      <div id="panelLog" class="hidden"></div>
      <div id="panelHelp" class="hidden"></div>
    </main>
  </div>
  <input type="file" id="importFile" accept=".json" class="hidden" />
  <script>
    const STORAGE_KEY = "{STORAGE_KEY}";
    const PLAN_DAYS = {data_json};
    const PLAN_META = {meta_json};
    const TRACK_DEFS = {tracks_json};

    let records = {{}};
    let currentId = PLAN_DAYS[0]?.id || "D1";
    let saveTimer = null;
    let activeTab = "plan";

    function loadRecords() {{
      try {{
        const raw = localStorage.getItem(STORAGE_KEY);
        if (raw) records = JSON.parse(raw);
      }} catch (e) {{ console.error(e); }}
    }}

    function defaultRecord(dayId) {{
      const day = PLAN_DAYS.find(d => d.id === dayId);
      const checks = {{}};
      (day?.checklistItems || []).forEach((_, i) => {{ checks[i] = false; }});
      return {{
        status: "pending",
        hours: "",
        checklist: checks,
        journal: "",
        review: "",
        note: "",
        updatedAt: ""
      }};
    }}

    function getRecord(dayId) {{
      if (!records[dayId]) records[dayId] = defaultRecord(dayId);
      return records[dayId];
    }}

    function scheduleSave() {{
      clearTimeout(saveTimer);
      document.getElementById("saveStatus").textContent = "保存中…";
      document.getElementById("saveStatus").classList.remove("saved");
      saveTimer = setTimeout(() => {{
        records[currentId].updatedAt = new Date().toISOString();
        localStorage.setItem(STORAGE_KEY, JSON.stringify(records));
        const t = new Date().toLocaleTimeString("zh-CN", {{ hour: "2-digit", minute: "2-digit", second: "2-digit" }});
        const el = document.getElementById("saveStatus");
        el.textContent = "已自动保存 " + t;
        el.classList.add("saved");
        renderDayList();
        updateProgress();
        if (activeTab === "log") renderLogPanel();
      }}, 400);
    }}

    function todayId() {{
      const today = new Date().toISOString().slice(0, 10);
      const found = PLAN_DAYS.find(d => d["日期"] === today);
      return found ? found.id : PLAN_DAYS[0].id;
    }}

    function dayStatus(dayId) {{
      const r = records[dayId];
      if (!r) return "";
      if (r.status === "done") return "done";
      if (r.status === "doing" || r.hours || r.journal || r.review) return "partial";
      const checks = Object.values(r.checklist || {{}});
      if (checks.some(Boolean)) return "partial";
      return "";
    }}

    function sumHours(ids) {{
      return ids.reduce((s, id) => s + (parseFloat(records[id]?.hours) || 0), 0);
    }}

    function trackDoneCount(track) {{
      return track.dayIds.filter(id => records[id]?.status === "done").length;
    }}

    function updateProgress() {{
      const done = PLAN_DAYS.filter(d => records[d.id]?.status === "done").length;
      const pct = Math.round((done / PLAN_DAYS.length) * 100);
      const actualH = sumHours(PLAN_DAYS.map(d => d.id));
      document.getElementById("progressFill").style.width = pct + "%";
      document.getElementById("progressText").textContent =
        `已完成 ${{done}}/${{PLAN_DAYS.length}} 天（${{pct}}%）· 计划 ${{PLAN_META.totalHours}}h · 已记录 ${{actualH || 0}}h`;
      if (activeTab === "progress") renderProgressPanel();
    }}

    function linkify(text) {{
      if (!text) return "";
      const escaped = String(text).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
      return escaped.replace(/(https?:\\/\\/[^\\s|]+)/g, '<a href="$1" target="_blank" rel="noopener">$1</a>');
    }}

    function renderDayList() {{
      const q = document.getElementById("search").value.trim().toLowerCase();
      const container = document.getElementById("dayList");
      container.innerHTML = "";
      let lastWeek = "";
      PLAN_DAYS.forEach(day => {{
        const text = JSON.stringify(day).toLowerCase();
        if (q && !text.includes(q) && !day.id.toLowerCase().includes(q)) return;
        if (day["周次"] !== lastWeek) {{
          lastWeek = day["周次"];
          const wl = document.createElement("div");
          wl.className = "week-label";
          wl.textContent = lastWeek + " · " + (day["阶段"] || "");
          container.appendChild(wl);
        }}
        const el = document.createElement("div");
        el.className = "day-item" + (day.id === currentId ? " active" : "") + " " + dayStatus(day.id);
        el.dataset.id = day.id;
        el.innerHTML = `
          <span class="day-dot"></span>
          <div>
            <div><span class="day-id">${{day.id}}</span> · ${{day["日期"]}} · ${{day["星期"]}}</div>
            <div class="day-meta">${{day["模块"]}} — ${{String(day["今日目标"]).slice(0, 18)}}…</div>
          </div>`;
        el.onclick = () => selectDay(day.id);
        container.appendChild(el);
      }});
    }}

    function selectDay(id) {{
      currentId = id;
      renderDayList();
      renderPlanPanel();
    }}

    function renderPlanPanel() {{
      const day = PLAN_DAYS.find(d => d.id === currentId);
      if (!day) return;
      const r = getRecord(currentId);
      const checks = day.checklistItems || [];
      const checkHtml = checks.length ? checks.map((item, i) => `
        <li>
          <input type="checkbox" data-check="${{i}}" ${{r.checklist?.[i] ? "checked" : ""}} />
          <span>${{item.replace(/^☐\\s*/, "")}}</span>
        </li>`).join("") : `<li class="field-value">${{day["验收清单"] || "—"}}</li>`;

      document.getElementById("panelPlan").innerHTML = `
        <div class="hero-title">${{day.id}} · ${{day["日期"]}}（${{day["星期"]}}）</div>
        <div class="hero-sub">${{day["阶段"]}} / ${{day["周次"]}} / ${{day["模块"]}}</div>
        <div class="tags">
          <span class="tag">建议 ${{day["建议学时"]}} 小时</span>
          <span class="tag">${{day["代码文件夹"] || "—"}}</span>
        </div>

        <div class="card">
          <h2>今日目标 & 交付物</h2>
          <p class="field-value"><strong>目标：</strong>${{day["今日目标"]}}</p>
          <p class="field-value" style="margin-top:10px"><strong>交付物：</strong>${{day["交付物（必达）"]}}</p>
        </div>

        <div class="grid-2">
          <div class="card">
            <h2>上午 / 下午</h2>
            <p class="field-label">上午任务</p>
            <p class="field-value">${{day["上午任务"]}}</p>
            <p class="field-label" style="margin-top:12px">下午任务</p>
            <p class="field-value">${{day["下午任务"]}}</p>
            <p class="field-label" style="margin-top:12px">晚上复盘提示</p>
            <p class="field-value">${{day["晚上复盘提示"]}}</p>
          </div>
          <div class="card">
            <h2>参考资料</h2>
            <p class="field-label">今日阅读</p>
            <p class="field-value">${{day["今日阅读章节"]}}</p>
            <p class="field-label">主文档</p>
            <p class="field-value">${{linkify(day["主文档链接"])}}（${{day["主文档（名称）"]}}）</p>
            <p class="field-label">教程 / 视频</p>
            <p class="field-value">${{linkify(day["入门教程/视频"])}}</p>
            <p class="field-label">怎么创建</p>
            <p class="field-value">${{day["怎么创建"]}}</p>
          </div>
        </div>

        <div class="card record-section">
          <h2>我的学习记录 <span style="font-size:12px;color:var(--muted);font-weight:normal">（修改后自动保存）</span></h2>
          <label>学习状态</label>
          <select id="recStatus">
            <option value="pending" ${{r.status==="pending"?"selected":""}}>未开始</option>
            <option value="doing" ${{r.status==="doing"?"selected":""}}>进行中</option>
            <option value="done" ${{r.status==="done"?"selected":""}}>已完成</option>
          </select>
          <label>实际学时（小时）</label>
          <input type="number" id="recHours" min="0" max="24" step="0.5" value="${{r.hours}}" placeholder="例如 6" />
          <label>验收清单（勾选即记录）</label>
          <ul class="checklist" id="recChecklist">${{checkHtml}}</ul>
          <label>学习日记（今天学了什么、代码/问题）</label>
          <textarea id="recJournal" placeholder="自由记录…">${{escapeHtml(r.journal)}}</textarea>
          <label>今晚复盘（学了啥 / 卡在哪 / 明天干啥）</label>
          <textarea id="recReview" placeholder="3句话复盘…">${{escapeHtml(r.review)}}</textarea>
          <label>其它备注</label>
          <textarea id="recNote" style="min-height:60px" placeholder="可选">${{escapeHtml(r.note)}}</textarea>
          ${{r.updatedAt ? `<p class="field-label" style="margin-top:12px">上次保存：${{new Date(r.updatedAt).toLocaleString("zh-CN")}}</p>` : ""}}
        </div>
      `;
      bindRecordEvents();
    }}

    function escapeHtml(s) {{
      return String(s||"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
    }}

    function bindRecordEvents() {{
      const r = getRecord(currentId);
      const bind = (id, key, isCheck) => {{
        const el = document.getElementById(id);
        if (!el) return;
        const evt = isCheck ? "change" : "input";
        el.addEventListener(evt, () => {{
          if (isCheck) {{
            if (!r.checklist) r.checklist = {{}};
            document.querySelectorAll("#recChecklist input[data-check]").forEach(cb => {{
              r.checklist[cb.dataset.check] = cb.checked;
            }});
          }} else {{
            r[key] = el.value;
          }}
          scheduleSave();
        }});
      }};
      bind("recStatus", "status");
      bind("recHours", "hours");
      bind("recJournal", "journal");
      bind("recReview", "review");
      bind("recNote", "note");
      document.querySelectorAll("#recChecklist input[data-check]").forEach(cb => {{
        cb.addEventListener("change", () => {{
          if (!r.checklist) r.checklist = {{}};
          r.checklist[cb.dataset.check] = cb.checked;
          scheduleSave();
        }});
      }});
      document.getElementById("recStatus")?.addEventListener("change", e => {{
        r.status = e.target.value;
        scheduleSave();
      }});
    }}

    function renderLogPanel() {{
      const rows = PLAN_DAYS.map(day => {{
        const r = records[day.id] || {{}};
        const preview = (r.journal || r.review || r.note || "").slice(0, 40);
        const statusMap = {{ pending: "未开始", doing: "进行中", done: "已完成" }};
        return `<tr data-goto="${{day.id}}" style="cursor:pointer">
          <td>${{day.id}}</td>
          <td>${{day["日期"]}}</td>
          <td>${{statusMap[r.status] || "未开始"}}</td>
          <td>${{r.hours || "—"}}</td>
          <td>${{preview || "—"}}${{preview.length >= 40 ? "…" : ""}}</td>
        </tr>`;
      }}).join("");
      document.getElementById("panelLog").innerHTML = `
        <div class="card">
          <h2>全部学习记录</h2>
          <p class="field-value" style="color:var(--muted);margin-bottom:16px">点击某行可跳转到该日。数据保存在本机浏览器，可「导出记录」备份。</p>
          <table class="log-table">
            <thead><tr><th>序号</th><th>日期</th><th>状态</th><th>学时</th><th>摘要</th></tr></thead>
            <tbody>${{rows}}</tbody>
          </table>
        </div>`;
      document.querySelectorAll("#panelLog tr[data-goto]").forEach(tr => {{
        tr.onclick = () => {{
          selectDay(tr.dataset.goto);
          setTab("plan");
        }};
      }});
    }}

    function renderProgressPanel() {{
      const today = new Date().toISOString().slice(0, 10);
      const todayDay = PLAN_DAYS.find(d => d["日期"] === today);
      const scheduleRows = PLAN_META.dailySchedule.map(r =>
        `<tr><td>${{r[0]}}</td><td>${{r[1]}}</td><td>${{r[2]}}</td></tr>`
      ).join("");
      const trackRows = TRACK_DEFS.map(track => {{
        const done = trackDoneCount(track);
        const total = track.dayIds.length;
        const tpct = total ? Math.round(done / total * 100) : 0;
        const ah = sumHours(track.dayIds);
        const status = done >= total ? "done" : (today >= track.dateStart && today <= track.dateEnd ? "active" : "");
        return `
          <div class="track-row">
            <div class="track-head">
              <span class="track-name">${{track.name}}</span>
              <span class="track-meta">${{track.dateStart}} ~ ${{track.dateEnd}} · ${{track.weeks}} · 计划 ${{track.hours}}h</span>
            </div>
            <div class="track-bar"><div class="track-fill" style="width:${{tpct}}%"></div></div>
            <div class="track-meta">${{done}}/${{total}} 天（${{tpct}}%）· 已记录 ${{ah || 0}}h · 交付：${{track.deliver}}</div>
          </div>`;
      }}).join("");
      const msItems = PLAN_META.milestones.map(ms => {{
        const cls = today >= ms.date ? "done" : (todayDay && todayDay.id >= ms.dayId ? "done" : "future");
        const r = records[ms.dayId];
        const hit = r?.status === "done" ? " ✓" : "";
        return `<div class="timeline-item ${{cls}}"><strong>${{ms.id}}</strong> · ${{ms.date}} · ${{ms.label}}${{hit}}</div>`;
      }}).join("");
      document.getElementById("panelProgress").innerHTML = `
        <div class="card">
          <h2>总时间与进程</h2>
          <p class="field-value"><strong>周期：</strong>${{PLAN_META.startDate}} ~ ${{PLAN_META.endDate}}（${{PLAN_META.totalDays}} 个工作日 / ${{PLAN_META.weeksCount}} 内容周 · 约 ${{PLAN_META.calendarWeeks}} 个自然周）</p>
          <p class="field-value"><strong>总计划学时：</strong>${{PLAN_META.totalHours}} 小时（每周 ${{PLAN_META.weeklyHours}}h = 6×${{PLAN_META.weekdayHours}}h + 1×${{PLAN_META.reviewDayHours}}h · 仅工作日学习 · 周末休息）</p>
          <p class="field-value"><strong>今天（${{today}}）：</strong>${{todayDay ? todayDay.id + " · " + todayDay["阶段"] + " · " + todayDay["今日目标"] : "不在计划期内"}}</p>
        </div>
        <div class="grid-2">
          <div class="card">
            <h2>每日时间分配（固定）</h2>
            <table class="log-table">
              <thead><tr><th>时段</th><th>时长</th><th>做什么</th></tr></thead>
              <tbody>${{scheduleRows}}</tbody>
            </table>
          </div>
          <div class="card">
            <h2>里程碑节点</h2>
            <div class="timeline">${{msItems}}</div>
          </div>
        </div>
        <div class="card">
          <h2>各方向学习进度（按「学习状态=已完成」统计）</h2>
          ${{trackRows}}
        </div>`;
    }}

    function renderHelpPanel() {{
      document.getElementById("panelHelp").innerHTML = `
        <div class="card">
          <h2>怎么用</h2>
          <ol class="field-value" style="padding-left:20px">
            <li>左侧点击某一天，右侧显示当日学习内容。</li>
            <li>在「我的学习记录」填写状态、学时、日记、复盘——<strong>自动保存</strong>到本机。</li>
            <li>点「今天」快速跳到计划表里今天的日期。</li>
            <li>「时间与进度」查看总学时、各方向进度条、里程碑。</li>
            <li>「学习记录总览」查看全部打卡情况。</li>
            <li>定期点「导出记录」备份 JSON，换电脑可用「导入记录」恢复。</li>
          </ol>
          <p class="field-value" style="margin-top:12px"><strong>准确时间：</strong>共 ${{PLAN_META.totalHours}} 小时 / ${{PLAN_META.totalDays}} 个工作日；仅周一至周五学习（每天 7h，每周最后一个工作日验收 6h），周末休息。</p>
          <p class="field-value" style="margin-top:16px;color:var(--warn)">注意：清除浏览器数据会丢失记录，请务必偶尔导出备份。</p>
          <p class="field-value">代码目录：<code>e:\\三个月速成\\learning-code</code></p>
        </div>`;
    }}

    function setTab(tab) {{
      activeTab = tab;
      document.querySelectorAll(".tab").forEach(t => t.classList.toggle("active", t.dataset.tab === tab));
      document.getElementById("panelPlan").classList.toggle("hidden", tab !== "plan");
      document.getElementById("panelProgress").classList.toggle("hidden", tab !== "progress");
      document.getElementById("panelLog").classList.toggle("hidden", tab !== "log");
      document.getElementById("panelHelp").classList.toggle("hidden", tab !== "help");
      if (tab === "progress") renderProgressPanel();
      if (tab === "log") renderLogPanel();
      if (tab === "help") renderHelpPanel();
    }}

    document.querySelectorAll(".tab").forEach(t => {{
      t.onclick = () => setTab(t.dataset.tab);
    }});
    document.getElementById("btnToday").onclick = () => selectDay(todayId());
    document.getElementById("search").oninput = renderDayList;
    document.getElementById("btnExport").onclick = () => {{
      const blob = new Blob([JSON.stringify({{ version: 1, exportedAt: new Date().toISOString(), records }}, null, 2)], {{ type: "application/json" }});
      const a = document.createElement("a");
      a.href = URL.createObjectURL(blob);
      a.download = "七方向学习记录_" + new Date().toISOString().slice(0,10) + ".json";
      a.click();
    }};
    document.getElementById("btnImport").onclick = () => document.getElementById("importFile").click();
    document.getElementById("importFile").onchange = e => {{
      const file = e.target.files[0];
      if (!file) return;
      const reader = new FileReader();
      reader.onload = () => {{
        try {{
          const data = JSON.parse(reader.result);
          records = data.records || data;
          localStorage.setItem(STORAGE_KEY, JSON.stringify(records));
          alert("导入成功！");
          renderDayList();
          updateProgress();
          renderPlanPanel();
          if (activeTab === "log") renderLogPanel();
          if (activeTab === "progress") renderProgressPanel();
        }} catch (err) {{ alert("导入失败：" + err.message); }}
      }};
      reader.readAsText(file);
      e.target.value = "";
    }};

    loadRecords();
    PLAN_DAYS.forEach(d => getRecord(d.id));
    const urlDay = new URLSearchParams(location.search).get("day");
    if (urlDay && PLAN_DAYS.some(d => d.id === urlDay)) currentId = urlDay;
    else if (PLAN_DAYS.some(d => d.id === todayId())) currentId = todayId();
    renderDayList();
    renderPlanPanel();
    renderProgressPanel();
    renderHelpPanel();
    updateProgress();
  </script>
</body>
</html>"""


def main():
    days = load_days()
    html = build_html(days)
    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"已生成: {HTML_OUT}")
    print(f"共 {len(days)} 天")


if __name__ == "__main__":
    main()
