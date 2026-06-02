# -*- coding: utf-8 -*-
"""将计划日期改为从 2026-06-01 起的 66 个工作日（仅周一至周五，8/31 结业）"""

import os
from datetime import date
from openpyxl import load_workbook

from rebuild_multi_track_plan import START, WEEKDAY_CN, learning_date, MILESTONE_WEEKS

DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DIR, "三个月全栈入门计划.xlsx")


def main():
    wb = load_workbook(XLSX)
    ws = wb["每日学习计划"]
    headers = [c.value for c in ws[1]]
    date_col = headers.index("日期") + 1
    week_col = headers.index("星期") + 1
    week_key_col = headers.index("周次") + 1
    id_col = 1

    week_dates = {}
    row = 2
    i = 0
    while ws.cell(row, id_col).value:
        d = learning_date(START, i)
        ws.cell(row, date_col, d.strftime("%Y-%m-%d"))
        ws.cell(row, week_col, WEEKDAY_CN[d.weekday()])
        wk = ws.cell(row, week_key_col).value
        week_dates.setdefault(wk, []).append(d.strftime("%Y-%m-%d"))
        row += 1
        i += 1

    if "周目标与参考书" in wb.sheetnames:
        ws2 = wb["周目标与参考书"]
        for r in range(2, ws2.max_row + 1):
            wk = ws2.cell(r, 1).value
            if wk in week_dates:
                dr = week_dates[wk]
                ws2.cell(r, 2, f"{dr[0]} ~ {dr[-1]}")

    if "里程碑" in wb.sheetnames:
        ws3 = wb["里程碑"]
        id_by_week = {}
        for r in range(2, ws.max_row + 1):
            wk = ws.cell(r, week_key_col).value
            id_by_week[wk] = ws.cell(r, id_col).value
        for r in range(2, ws3.max_row + 1):
            node = ws3.cell(r, 1).value
            if node in MILESTONE_WEEKS:
                wk = MILESTONE_WEEKS[node]
                idx = int(str(id_by_week[wk]).lstrip("D")) - 1
                d = learning_date(START, idx)
                ws3.cell(r, 2, d.strftime("%Y-%m-%d"))

    wb.save(XLSX)
    end = learning_date(START, i - 1)
    print(f"已更新 {i} 个工作日: {START} ~ {end}（周末休息）")


if __name__ == "__main__":
    main()
