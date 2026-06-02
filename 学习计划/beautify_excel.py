# -*- coding: utf-8 -*-
"""美化「三个月全栈入门计划.xlsx」样式：分阶段配色、周次区分、表头与边框"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DIR, "三个月全栈入门计划.xlsx")

# 阶段底色（主色 / 斑马纹）
PHASE_FILL = {
    "第1月·工具与读码": ("E8EAF6", "F3E5F5"),
    "第1月·H5验收向": ("E8F5E9", "F1F8E9"),
    "第1月·后端Java": ("E3F2FD", "E8EAF6"),
    "第1月·后端API": ("DCEBFF", "EDF4FF"),
    "第2月·Android": ("BBDEFB", "E3F2FD"),
    "第2月·Android进阶": ("90CAF9", "E1F5FE"),
    "第2月·Cocos": ("FFE0B2", "FFF3E0"),
    "第2月·Unity3D": ("D1C4E9", "EDE7F6"),
    "第3月·后端进阶": ("C5CAE9", "E8EAF6"),
    "第3月·多端联调": ("FFE8D6", "FFF3EB"),
    "第3月·运维认知": ("B2DFDB", "E0F2F1"),
    "第3月·结业": ("FFF4CC", "FFFBEB"),
    # 兼容旧阶段名
    "第1月·前端": ("DFF5E4", "F0FAF2"),
    "第2月·后端": ("DCEBFF", "EDF4FF"),
    "第3月·联调": ("FFE8D6", "FFF3EB"),
    "第3月·上线": ("FFF4CC", "FFFBEB"),
}

# 周次左边框强调色（12 周）
WEEK_ACCENT = {
    "W1": "43A047", "W2": "66BB6A", "W3": "2E7D32", "W4": "1B5E20",
    "W5": "1E88E5", "W6": "42A5F5", "W7": "1565C0", "W8": "0D47A1",
    "W9": "FB8C00", "W10": "F57C00", "W11": "EF6C00", "W12": "E65100",
}

HEADER_COLORS = {
    "default": "1A365D",
    "week": "2C5282",
    "milestone": "553C9A",
    "tools": "276749",
    "links": "234E52",
    "help": "744210",
}

THIN = Side(style="thin", color="B8C5D6")
MEDIUM = Side(style="medium", color="7A8FA6")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def style_header_row(ws, row=1, bg="1A365D", max_col=None):
    max_col = max_col or ws.max_column
    for col in range(1, max_col + 1):
        c = ws.cell(row, col)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        if isinstance(col, str):
            ws.column_dimensions[col].width = w
        else:
            ws.column_dimensions[get_column_letter(col)].width = w


def beautify_daily_plan(ws):
    max_row = ws.max_row
    max_col = ws.max_column
  # 找列号
    headers = {ws.cell(1, c).value: c for c in range(1, max_col + 1)}
    col_phase = headers.get("阶段", 5)
    col_week = headers.get("周次", 4)
    col_done = headers.get("完成", 21)

    style_header_row(ws, bg=HEADER_COLORS["default"], max_col=max_col)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    widths = {
        1: 6, 2: 12, 3: 6, 4: 6, 5: 12, 6: 9,
        7: 24, 8: 22, 9: 16, 10: 20, 11: 14, 12: 28,
        13: 24, 14: 26, 15: 14, 16: 22, 17: 22, 18: 14,
        19: 22, 20: 9, 21: 7, 22: 9, 23: 14,
    }
    set_col_widths(ws, widths)

    dv = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
    ws.add_data_validation(dv)
    if col_done:
        dv.add(f"{get_column_letter(col_done)}2:{get_column_letter(col_done)}{max_row}")

    last_week = None
    for row in range(2, max_row + 1):
        phase = str(ws.cell(row, col_phase).value or "")
        week = str(ws.cell(row, col_week).value or "")
        main_c, alt_c = PHASE_FILL.get(phase, ("F5F7FA", "FFFFFF"))
        use_alt = (row % 2 == 0)
        row_fill = fill(alt_c if use_alt else main_c)

        # 新的一周：略加深行高 + 顶部粗线
        if week != last_week:
            ws.row_dimensions[row].height = 28
            top_border = Border(
                left=THIN, right=THIN,
                top=MEDIUM, bottom=THIN,
            )
            last_week = week
        else:
            ws.row_dimensions[row].height = 52
            top_border = BORDER_ALL

        accent = WEEK_ACCENT.get(week, "607D8B")
        left_accent = Border(
            left=Side(style="thick", color=accent),
            right=THIN, top=top_border.top, bottom=THIN,
        )

        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = left_accent if col == 1 else Border(
                left=THIN, right=THIN, top=top_border.top, bottom=THIN
            )
            # 关键列加粗
            if col in (1, 4, 7):
                cell.font = Font(size=10, bold=(col in (1, 4)))
            else:
                cell.font = Font(size=10)
            # 完成列居中
            if col == col_done:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value == "是":
                    cell.fill = fill("C8E6C9")
                    cell.font = Font(bold=True, color="1B5E20")
            # 链接列
            if col == headers.get("主文档链接", 0):
                v = cell.value
                if v and str(v).startswith("http"):
                    cell.font = Font(size=10, color="0563C1", underline="single")
                    cell.hyperlink = str(v)

    # 阶段图例（右上角备注行下方插一行说明 — 用第 max_row+2 若空间紧则跳过）
    note_row = max_row + 2
    ws.cell(note_row, 1, "配色说明").font = Font(bold=True, size=10)
    legend = [
        ("W1 工具读码", "E8EAF6"), ("W2 H5验收", "E8F5E9"),
        ("后端 W3-W4/W9", "DCEBFF"), ("Android W5-W6", "BBDEFB"),
        ("Cocos W7", "FFE0B2"), ("Unity W8", "D1C4E9"),
        ("联调 W10", "FFE8D6"), ("运维 W11", "B2DFDB"), ("结业 W12", "FFF4CC"),
    ]
    for i, (label, color) in enumerate(legend):
        col = 2 + (i % 5) * 2
        row_off = i // 5
        c = ws.cell(note_row + row_off, col, label)
        c.fill = fill(color)
        c.border = BORDER_ALL
        c.font = Font(size=9)


def beautify_table_sheet(ws, header_color="2C5282", zebra=False):
    max_row = ws.max_row
    max_col = ws.max_column
    style_header_row(ws, bg=header_color, max_col=max_col)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    for row in range(2, max_row + 1):
        if not any(ws.cell(row, c).value for c in range(1, max_col + 1)):
            continue
        bg = "F7FAFC" if zebra and row % 2 == 0 else "FFFFFF"
        # 周次列上色
        week_val = ws.cell(row, 1).value
        if week_val in WEEK_ACCENT:
            bg = PHASE_FILL.get(
                "第1月·前端" if week_val in ("W1", "W2", "W3", "W4") else
                "第2月·后端" if week_val in ("W5", "W6", "W7", "W8") else
                "第3月·联调" if week_val in ("W9", "W10", "W11") else "第3月·上线",
                ("F5F7FA", "FFFFFF"),
            )[0]
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = fill(bg)
            cell.border = BORDER_ALL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
            if col == 1 and str(week_val).startswith("W"):
                cell.font = Font(bold=True, color=WEEK_ACCENT.get(week_val, "333333"))
            v = cell.value
            if v and str(v).startswith("http"):
                cell.font = Font(size=10, color="0563C1", underline="single")
                cell.hyperlink = str(v)
        ws.row_dimensions[row].height = 36


def beautify_usage(ws):
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 72
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row, 1).value
        if row == 1:
            ws.cell(row, 1).font = Font(bold=True, size=14, color="C05621")
        elif a and str(a).startswith("【"):
            ws.cell(row, 1).font = Font(bold=True, size=12, color="1A365D")
        for col in (1, 2):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")


def beautify_milestones(ws):
    beautify_table_sheet(ws, HEADER_COLORS["milestone"])
    for row in range(2, ws.max_row + 1):
        node = ws.cell(row, 1).value
        colors = {"M1": "DFF5E4", "M2": "DCEBFF", "M3": "FFE8D6"}
        if node in colors:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = fill(colors[node])


def main():
    wb = load_workbook(XLSX)

    if "每日学习计划" in wb.sheetnames:
        beautify_daily_plan(wb["每日学习计划"])

    sheet_styles = {
        "周目标与参考书": HEADER_COLORS["week"],
        "三个月总览": HEADER_COLORS["default"],
        "怎么学": HEADER_COLORS["default"],
        "工具安装": HEADER_COLORS["tools"],
        "链接速查": HEADER_COLORS["links"],
        "书单与视频": HEADER_COLORS["help"],
        "7天极简入门": HEADER_COLORS["help"],
        "常见问题": HEADER_COLORS["default"],
        "H5学习目标": "2E7D32",
        "运维学习目标": "00695C",
        "学习时间与进度": "4527A0",
        "里程碑": None,
    }
    for name, color in sheet_styles.items():
        if name not in wb.sheetnames:
            continue
        if name == "里程碑":
            beautify_milestones(wb[name])
        else:
            beautify_table_sheet(wb[name], color, zebra=True)

    if "使用说明" in wb.sheetnames:
        beautify_usage(wb["使用说明"])

    # 目标/进度 sheet 列宽
    for name, widths in {
        "H5学习目标": {"A": 8, "B": 52, "C": 14, "D": 16},
        "运维学习目标": {"A": 8, "B": 52, "C": 14, "D": 16},
        "学习时间与进度": {"A": 18, "B": 22, "C": 12, "D": 10, "E": 12, "F": 28},
    }.items():
        if name in wb.sheetnames:
            ws = wb[name]
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

    tab_colors = {
        "使用说明": "FFE082",
        "每日学习计划": "4CAF50",
        "H5学习目标": "66BB6A",
        "运维学习目标": "00897B",
        "学习时间与进度": "7E57C2",
        "怎么学": "64B5F6",
        "周目标与参考书": "42A5F5",
        "三个月总览": "7986CB",
        "里程碑": "AB47BC",
        "工具安装": "66BB6A",
        "链接速查": "26A69A",
        "书单与视频": "8D6E63",
        "7天极简入门": "FFA726",
        "常见问题": "90A4AE",
    }
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.save(XLSX)
    print(f"已美化: {XLSX}")


if __name__ == "__main__":
    main()
